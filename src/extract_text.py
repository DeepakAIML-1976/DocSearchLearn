# extract_text.py
import fitz  # PyMuPDF
import docx
import openpyxl
import pytesseract
from PIL import Image

def extract_pdf_text(path):
    text = ""
    with fitz.open(path) as pdf:
        for page in pdf:
            text += page.get_text()
    return text

def extract_docx_text(path):
    doc = docx.Document(path)
    return "\n".join([p.text for p in doc.paragraphs])

def extract_excel_text(path):
    wb = openpyxl.load_workbook(path)
    text = ""
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            text += " ".join([str(cell) if cell else "" for cell in row]) + "\n"
    return text

def extract_image_text(path):
    img = Image.open(path)
    return pytesseract.image_to_string(img)
